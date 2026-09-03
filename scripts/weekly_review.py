"""Weekly model review from frozen prediction snapshots and next-day outcomes."""
from __future__ import annotations

import json
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def load_json(path: Path):
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None


def collect_validation() -> list[dict]:
    rows: list[dict] = []
    for path in sorted(Path("data/validation").glob("*.json")):
        if path.name == "latest_validation.json":
            continue
        data = load_json(path)
        if isinstance(data, list):
            rows.extend(data)
    return rows


def _score_band(score: float) -> str:
    if score < 64:
        return "<64"
    if score < 68:
        return "64-67.9"
    if score < 72:
        return "68-71.9"
    if score < 76:
        return "72-75.9"
    return "76+"


def _risk_tags(row: dict) -> list[str]:
    raw = str(row.get("失败风险标签") or "")
    return [x.strip() for x in raw.split("；") if x.strip()]


def main() -> None:
    rows = collect_validation()
    if not rows:
        Path("reports").mkdir(parents=True, exist_ok=True)
        Path("reports/weekly_model_review.json").write_text(json.dumps({"samples": 0, "status": "样本不足"}, ensure_ascii=False, indent=2), encoding="utf-8")
        print("weekly_review_samples=0")
        return

    total = len(rows)
    wins = sum(r.get("次日上涨") == "是" for r in rows)
    failures = [r for r in rows if r.get("结果") == "失败"]
    strong = sum(r.get("次日≥3%") == "是" for r in rows)
    avg_ret = sum(float(r.get("次日收益%", 0) or 0) for r in rows) / total

    by_band = defaultdict(list)
    for r in rows:
        try:
            score = float(r.get("Score"))
        except (TypeError, ValueError):
            continue
        by_band[_score_band(score)].append(r)

    tag_total = Counter()
    tag_failures = Counter()
    tag_returns = defaultdict(list)
    for r in failures:
        for tag in _risk_tags(r):
            tag_failures[tag] += 1
            tag_returns[tag].append(float(r.get("次日收益%", 0) or 0))
    for r in rows:
        for tag in _risk_tags(r):
            tag_total[tag] += 1

    wb = Workbook()
    ws = wb.active
    ws.title = "周度总览"
    ws.append(["指标", "值"])
    ws.append(["验证样本", total])
    ws.append(["次日上涨", wins])
    ws.append(["次日下跌", total - wins])
    ws.append(["次日胜率", round(wins / total, 4)])
    ws.append(["重点观察失败数", len([r for r in failures if r.get("重点观察") == "是"])])
    ws.append(["次日≥3%", strong])
    ws.append(["平均次日收益%", round(avg_ret, 4)])

    wb2 = wb.create_sheet("Score分段")
    wb2.append(["Score区间", "样本", "胜率", "平均收益%", "失败数"])
    for band in ["<64", "64-67.9", "68-71.9", "72-75.9", "76+"]:
        subset = by_band.get(band, [])
        n = len(subset)
        if not n:
            wb2.append([band, 0, None, None, 0])
            continue
        w = sum(x.get("次日上涨") == "是" for x in subset)
        a = sum(float(x.get("次日收益%", 0) or 0) for x in subset) / n
        wb2.append([band, n, round(w / n, 4), round(a, 4), n - w])

    wb3 = wb.create_sheet("失败归因")
    wb3.append(["风险标签", "命中样本", "其中失败", "失败占比", "命中样本平均收益%", "说明"])
    for tag in sorted(set(tag_total) | set(tag_failures)):
        n = tag_total[tag]
        f = tag_failures[tag]
        avg = sum(tag_returns[tag]) / f if f else None
        wb3.append([tag, n, f, round(f / len(failures), 4) if failures else None, round(avg, 4) if avg is not None else None, "预先定义的风险标签，不代表因果关系；一只股票可同时命中多个标签"])

    wb4 = wb.create_sheet("失败明细")
    failure_headers = ["预测日期", "验证日期", "代码", "名称", "Score", "观察", "重点观察", "预测等级", "昨日收盘", "次日收盘", "次日收益%", "失败风险标签"]
    wb4.append(failure_headers)
    for r in failures:
        wb4.append([r.get(h) for h in failure_headers])

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for c in sheet[1]:
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="D9EAF7")
            c.alignment = Alignment(horizontal="center")
        for column_cells in sheet.columns:
            letter = column_cells[0].column_letter
            width = min(42, max(len(str(c.value or "")) for c in column_cells) + 2)
            sheet.column_dimensions[letter].width = width

    out = Path("reports") / "weekly_model_review.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    summary = {
        "samples": total,
        "win_rate": round(wins / total, 4),
        "avg_return_pct": round(avg_ret, 4),
        "strong_up_count": strong,
        "failure_count": len(failures),
        "focus_samples": sum(r.get("重点观察") == "是" for r in rows),
        "focus_win_rate": round(sum(r.get("重点观察") == "是" and r.get("次日上涨") == "是" for r in rows) / max(1, sum(r.get("重点观察") == "是" for r in rows)), 4),
        "failure_tag_counts": dict(tag_failures),
    }
    Path("reports/weekly_model_review.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"weekly_review_samples={total} win_rate={summary['win_rate']} avg_return_pct={summary['avg_return_pct']} failures={len(failures)}")


if __name__ == "__main__":
    main()
