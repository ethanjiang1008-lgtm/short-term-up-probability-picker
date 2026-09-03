"""Daily frozen prediction snapshots, next-day validation, and Excel reports."""
from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path

from labels import next_day_labels
from market_data_sina import fetch_klines_parallel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


def load_json(path: Path):
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None


def save_daily_snapshot(rows: list[dict]) -> Path:
    day = str(rows[0]["date"])
    out = Path("data/predictions") / f"{day}_tail.json"
    out.parent.mkdir(parents=True, exist_ok=True)
    payload = {"date": day, "captured_at_utc": datetime.now(timezone.utc).isoformat(), "rows": rows}
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return out


def _flat(row: dict) -> dict:
    t, c, e = row.get("technical") or {}, row.get("components") or {}, row.get("evidence") or {}
    return {
        "日期": row.get("date"), "代码": row.get("code"), "名称": row.get("name"), "收盘价": row.get("today_close"),
        "涨幅%": row.get("today_change_pct"), "换手率%": row.get("turnover_rate_pct"), "量比代理": row.get("volume_ratio"),
        "Score": row.get("score"), "等级": row.get("grade"), "观察": row.get("observation"),
        "重点观察": "是" if row.get("is_focus") else "否", "涨停观察": "是" if row.get("is_limit_up") else "否",
        "MA5": t.get("ma5"), "MA10": t.get("ma10"), "MA20": t.get("ma20"), "MA60": t.get("ma60"),
        "MA5向上": "是" if t.get("ma5_rising") else "否", "站上MA20": "是" if t.get("close_above_ma20") else "否",
        "均线多头": "是" if t.get("ma_bull_alignment") else "否", "连续上涨": t.get("consecutive_up_days"),
        "5日收益": e.get("ret_5d"), "20日收益": e.get("ret_20d"), "相对量能": e.get("relative_volume_5d_vs_20d"),
        "尾盘加速": e.get("tail_acceleration"), "基础强度": c.get("base_strength"), "尾盘强度": c.get("close_strength"),
        "板块强度": c.get("sector_strength"), "事件强度": c.get("event_strength"), "市场环境": c.get("market_environment"),
        "判断理由": row.get("focus_reason"),
    }


def _style_sheet(ws, fill="D9EAF7"):
    ws.freeze_panes = "A2"
    if ws.max_row >= 1:
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.alignment = Alignment(horizontal="center")
    for column_cells in ws.columns:
        letter = column_cells[0].column_letter
        max_len = min(32, max(len(str(c.value or "")) for c in column_cells) + 2)
        ws.column_dimensions[letter].width = max_len


def export_excel(rows: list[dict], day: str) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "尾盘预测"
    flat_rows = [_flat(r) for r in rows]
    if flat_rows:
        headers = list(flat_rows[0].keys())
        ws.append(headers)
        for item in flat_rows:
            ws.append([item.get(h) for h in headers])
    _style_sheet(ws)
    out = Path("reports") / f"{day}_tail_prediction.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def _bootstrap_existing_daily_candidates() -> None:
    pred_dir = Path("data/predictions")
    pred_dir.mkdir(parents=True, exist_ok=True)
    if list(pred_dir.glob("*_tail.json")):
        return
    rows = load_json(Path("data/daily_candidates.json"))
    if isinstance(rows, list) and rows:
        save_daily_snapshot(rows)
        export_excel(rows, str(rows[0].get("date")))
        print(f"bootstrapped_prediction_date={rows[0].get('date')} rows={len(rows)}")


def _failure_reasons(row: dict, next_day_return: float) -> list[str]:
    t = row.get("technical") or {}
    e = row.get("evidence") or {}
    market = row.get("market") or {}
    reasons: list[str] = []
    close = float(row.get("today_close") or 0)
    ma5 = float(t.get("ma5") or 0)
    ma20 = float(t.get("ma20") or 0)
    ret5 = float(e.get("ret_5d") or 0)
    ret20 = float(e.get("ret_20d") or 0)
    tail_accel = float(e.get("tail_acceleration") or 0)
    relvol = float(e.get("relative_volume_5d_vs_20d") or 0)
    volume_proxy = float(row.get("volume_ratio") or 0)
    turnover = float(row.get("turnover_rate_pct") or 0)
    up_days = int(t.get("consecutive_up_days") or 0)
    breadth = float(market.get("breadth") or 0)
    market_env = float((row.get("components") or {}).get("market_environment") or 0)
    if ret5 >= 0.15 or ret20 >= 0.25 or up_days >= 4 or (close > 0 and ma5 > 0 and close / ma5 - 1 >= 0.08):
        reasons.append("短期过热/乖离偏高")
    if tail_accel >= 0.05:
        reasons.append("尾盘加速过快")
    if relvol >= 1.8 or volume_proxy >= 2.5:
        reasons.append("放量风险")
    if turnover >= 8.0:
        reasons.append("高换手风险")
    if breadth < 0.35 or market_env < 45:
        reasons.append("市场环境偏弱")
    if close > 0 and ma20 > 0 and close / ma20 - 1 < 0.05 and not bool(t.get("ma_bull_alignment")):
        reasons.append("趋势确认不足")
    if next_day_return < 0 and not reasons:
        reasons.append("未命中预设风险标签")
    return reasons


def validate_previous_day() -> tuple[str | None, list[dict]]:
    _bootstrap_existing_daily_candidates()
    files = sorted(Path("data/predictions").glob("*_tail.json"))
    if not files:
        return None, []
    latest = load_json(files[-1]) or {}
    day = str(latest.get("date", ""))
    rows = latest.get("rows") or []
    if not day or not rows:
        return None, []
    codes = [str(r.get("code")) for r in rows if r.get("code")]
    bars_map, _ = fetch_klines_parallel(codes, count=10, workers=10)
    results = []
    for r in rows:
        code = str(r.get("code"))
        future = []
        for bar in bars_map.get(code, []):
            d = str(bar.get("date", bar.get("day", "")))[:10]
            try:
                close = float(bar.get("close", bar.get("c")))
            except (TypeError, ValueError):
                continue
            if d > day:
                future.append((d, close))
        if not future or not r.get("today_close"):
            continue
        next_day, next_close = sorted(future)[0]
        labels = next_day_labels(float(r["today_close"]), next_close)
        next_return = float(labels.get("next_day_return", 0) or 0)
        failure = next_return <= 0
        results.append({
            "预测日期": day, "验证日期": next_day, "代码": code, "名称": r.get("name"), "Score": r.get("score"),
            "观察": r.get("observation"), "重点观察": "是" if r.get("is_focus") else "否", "预测等级": r.get("grade"),
            "昨日收盘": r.get("today_close"), "次日收盘": next_close,
            "次日收益%": round(next_return * 100, 3), "次日上涨": "是" if labels.get("next_day_up") else "否",
            "次日≥3%": "是" if labels.get("next_day_strong_up") else "否", "次日涨停": "是" if labels.get("next_day_limit_up") else "否",
            "结果": "失败" if failure else "成功", "失败风险标签": "；".join(_failure_reasons(r, next_return)) if failure else "",
        })
    return day, results


def save_validation(day: str, results: list[dict]) -> Path:
    out = Path("data/validation") / f"{day}_validation.json"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
    summary = {
        "prediction_date": day,
        "validation_date": results[0].get("验证日期") if results else None,
        "samples": len(results),
        "wins": sum(r.get("次日上涨") == "是" for r in results),
        "losses": sum(r.get("次日上涨") == "否" for r in results),
        "win_rate": round(sum(r.get("次日上涨") == "是" for r in results) / len(results), 4) if results else None,
        "avg_return_pct": round(sum(float(r.get("次日收益%", 0) or 0) for r in results) / len(results), 4) if results else None,
        "strong_up_count": sum(r.get("次日≥3%") == "是" for r in results),
        "limit_up_count": sum(r.get("次日涨停") == "是" for r in results),
        "focus_samples": sum(r.get("重点观察") == "是" for r in results),
        "focus_wins": sum(r.get("重点观察") == "是" and r.get("次日上涨") == "是" for r in results),
        "focus_win_rate": round(sum(r.get("重点观察") == "是" and r.get("次日上涨") == "是" for r in results) / max(1, sum(r.get("重点观察") == "是" for r in results)), 4),
    }
    Path("data/validation/latest_validation.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return out


def append_validation_to_excel(day: str, results: list[dict]) -> Path:
    path = Path("reports") / f"{day}_tail_prediction.xlsx"
    if not path.exists() or not results:
        return path
    wb = load_workbook(path)
    if "次日验证" in wb.sheetnames:
        del wb["次日验证"]
    ws = wb.create_sheet("次日验证")
    headers = list(results[0].keys())
    ws.append(headers)
    for item in results:
        ws.append([item.get(h) for h in headers])
    _style_sheet(ws, "E2F0D9")
    wb.save(path)
    return path


def main_validate() -> None:
    day, results = validate_previous_day()
    if day:
        save_validation(day, results)
        append_validation_to_excel(day, results)
        print(f"validated_prediction_date={day} rows={len(results)}")
    else:
        print("validated_prediction_date=none rows=0")


if __name__ == "__main__":
    main_validate()
